"""
Excel文件处理工具
"""
import pandas as pd
from io import BytesIO
from datetime import datetime, timezone
from models import db
from models.user import User
from models.class_model import Class
from models.score import Score
from models.exam import Exam


class ExcelHandler:
    """Excel文件处理类"""
    
    @staticmethod
    def import_scores_from_excel(file, teacher_id):
        """
        从Excel文件导入成绩
        
        支持的列名（中英文均可）:
        - student_id / 学号 / 学生ID
        - class_id / 班级ID / 班级编号
        - subject / 科目 / 课程
        - score / 成绩 / 得分
        - total_score / 总分 / 满分
        - exam_id / 考试ID (可选)
        - comments / 备注 / 说明 (可选)
        
        Args:
            file: Flask上传的文件对象
            teacher_id: 教师ID
            
        Returns:
            dict: 包含success_count和errors的字典
        """
        errors = []
        success_count = 0
        
        try:
            # 读取Excel文件
            df = pd.read_excel(file)
            
            # 列名映射（支持中英文列名）
            column_mapping = {
                'student_id': ['student_id', '学号', '学生id', '学生ID', 'studentid', 'studentId', '学生编号'],
                'class_id': ['class_id', '班级id', '班级ID', '班级编号', 'classid', 'classId'],
                'subject': ['subject', '科目', '课程', '课程名称'],
                'score': ['score', '成绩', '得分', '分数'],
                'total_score': ['total_score', '总分', '满分', 'totalscore', 'totalScore'],
                'exam_id': ['exam_id', '考试id', '考试ID', 'examid', 'examId'],
                'comments': ['comments', '备注', '说明', 'comment']
            }
            
            # 标准化列名
            normalized_columns = {}
            for standard_name, possible_names in column_mapping.items():
                for col in df.columns:
                    col_str = str(col).strip()
                    if col_str in possible_names or col_str.lower() == standard_name.lower():
                        normalized_columns[standard_name] = col
                        break
            
            # 验证必需的列
            required_columns = ['student_id', 'class_id', 'subject', 'score', 'total_score']
            missing_columns = [col for col in required_columns if col not in normalized_columns]
            
            if missing_columns:
                errors.append(f'缺少必需的列: {", ".join(missing_columns)}。请确保Excel包含以下列之一：学号/student_id, 班级ID/class_id, 科目/subject, 成绩/score, 总分/total_score')
                return {'success_count': 0, 'errors': errors}
            
            # 处理每一行数据
            for index, row in df.iterrows():
                try:
                    # 获取数据（使用标准列名）
                    student_identifier = row.get(normalized_columns['student_id'])
                    class_id = row.get(normalized_columns['class_id'])
                    subject = row.get(normalized_columns['subject'])
                    score = row.get(normalized_columns['score'])
                    total_score = row.get(normalized_columns['total_score'])
                    exam_id = row.get(normalized_columns.get('exam_id'), None)
                    comments = row.get(normalized_columns.get('comments'), '')
                    
                    # 验证数据
                    if pd.isna(student_identifier) or pd.isna(class_id) or pd.isna(subject) or pd.isna(score) or pd.isna(total_score):
                        errors.append(f'第{index + 2}行: 必填字段不能为空')
                        continue
                    
                    # 验证班级是否属于该教师
                    class_ = Class.query.filter_by(
                        id=int(class_id),
                        teacher_id=teacher_id
                    ).first()
                    
                    if not class_:
                        errors.append(f'第{index + 2}行: 班级不存在或无权限访问 (class_id: {class_id})')
                        continue
                    
                    # 验证学生是否存在（支持通过ID或学号查找）
                    # 先获取学生的 role_id
                    try:
                        from database import Database
                        role_query = "SELECT role_id FROM roles WHERE role_name = 'student' LIMIT 1"
                        role_result = Database.execute_query(role_query, fetch_one=True)
                        student_role_id = role_result['role_id'] if role_result else None
                    except:
                        student_role_id = None
                    
                    # 尝试通过ID或学号查找学生
                    student = None
                    try:
                        # 先尝试作为ID查找
                        student_id_int = int(float(student_identifier))
                        query = User.query.filter_by(id=student_id_int)
                        if student_role_id:
                            query = query.filter_by(role_id=student_role_id)
                        student = query.first()
                    except:
                        pass
                    
                    # 如果通过ID没找到，尝试通过学号查找
                    if not student:
                        query = User.query.filter_by(system_account=str(student_identifier).strip())
                        if student_role_id:
                            query = query.filter_by(role_id=student_role_id)
                        student = query.first()
                    
                    if not student:
                        errors.append(f'第{index + 2}行: 学生不存在 (学号/ID: {student_identifier})')
                        continue
                    
                    student_id = student.id
                    
                    # 验证考试是否存在（如果提供了exam_id）
                    if exam_id and not pd.isna(exam_id):
                        exam = Exam.query.filter_by(id=int(exam_id)).first()
                        if not exam:
                            errors.append(f'第{index + 2}行: 考试不存在 (exam_id: {exam_id})')
                            continue
                    
                    # 创建成绩记录
                    # 使用当前时间（UTC时区）
                    current_time = datetime.now(timezone.utc)
                    
                    score_record = Score(
                        student_id=int(student_id),
                        class_id=int(class_id),
                        exam_id=int(exam_id) if exam_id and not pd.isna(exam_id) else None,
                        subject=str(subject),
                        score=float(score),
                        total_score=float(total_score),
                        type='imported',
                        recorded_by=teacher_id,
                        comments=str(comments) if comments and not pd.isna(comments) else None
                    )
                    
                    # 手动设置时间字段，确保使用当前时间而不是默认值
                    # 这样可以避免数据库默认值覆盖手动设置的值
                    score_record.recorded_at = current_time
                    score_record.created_at = current_time
                    score_record.updated_at = current_time
                    
                    print(f'[导入成绩] 第 {index + 2} 行 - 设置时间: {current_time} (UTC), recorded_at: {score_record.recorded_at}')
                    
                    db.session.add(score_record)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'第{index + 2}行: {str(e)}')
                    continue
            
            # 提交所有更改
            if success_count > 0:
                db.session.commit()
            
        except Exception as e:
            errors.append(f'文件处理错误: {str(e)}')
            db.session.rollback()
        
        return {
            'success_count': success_count,
            'errors': errors
        }
    
    @staticmethod
    def export_scores_to_excel(scores):
        """
        导出成绩到Excel文件
        
        Args:
            scores: 成绩数据列表（通常是查询结果）
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        try:
            # 将查询结果转换为DataFrame
            print(f'[ExcelHandler] 收到 {len(scores) if scores else 0} 条成绩记录')
            
            if not scores or len(scores) == 0:
                # 如果没有数据，创建空DataFrame
                print('[ExcelHandler] 没有数据，创建空DataFrame')
                df = pd.DataFrame(columns=[
                    '学生姓名', '学号', '班级名称', '考试标题', 
                    '科目', '得分', '总分', '百分比', '类型', '备注', '记录时间'
                ])
            else:
                # 转换数据
                print(f'[ExcelHandler] 开始转换数据，第一条记录类型: {type(scores[0])}')
                if isinstance(scores[0], dict):
                    print(f'[ExcelHandler] 第一条记录键名: {list(scores[0].keys())}')
                    print(f'[ExcelHandler] 第一条记录内容: {scores[0]}')
                
                data = []
                for idx, score in enumerate(scores):
                    try:
                        # ScoreService.get_scores_for_export 返回的字典键名是中文
                        if isinstance(score, dict):
                            # 直接使用字典中的数据
                            row_data = {
                                '学生姓名': score.get('学生姓名', ''),
                                '学号': score.get('学号', ''),
                                '班级名称': score.get('班级', ''),  # 注意：字典中是'班级'，Excel列名是'班级名称'
                                '考试标题': score.get('考试名称', ''),  # 注意：字典中是'考试名称'，Excel列名是'考试标题'
                                '科目': score.get('科目', ''),
                                '得分': score.get('成绩', ''),  # 注意：字典中是'成绩'，Excel列名是'得分'
                                '总分': score.get('总分', ''),
                                '百分比': score.get('百分比', ''),
                                '类型': score.get('录入方式', ''),  # 注意：字典中是'录入方式'，Excel列名是'类型'
                                '备注': score.get('备注', ''),
                                '记录时间': score.get('录入时间', '')  # 注意：字典中是'录入时间'，Excel列名是'记录时间'
                            }
                            data.append(row_data)
                            if idx == 0:
                                print(f'[ExcelHandler] 第一条转换后的数据: {row_data}')
                        else:
                            # 如果是对象，使用属性访问
                            row_data = {
                                '学生姓名': getattr(score, 'student_name', '') if hasattr(score, 'student_name') else '',
                                '学号': getattr(score, 'student_id', '') if hasattr(score, 'student_id') else '',
                                '班级名称': getattr(score, 'class_name', '') if hasattr(score, 'class_name') else '',
                                '考试标题': getattr(score, 'exam_title', '') if hasattr(score, 'exam_title') else '',
                                '科目': getattr(score, 'subject', '') if hasattr(score, 'subject') else '',
                                '得分': getattr(score, 'score', '') if hasattr(score, 'score') else '',
                                '总分': getattr(score, 'total_score', '') if hasattr(score, 'total_score') else '',
                                '百分比': f"{getattr(score, 'percentage', 0):.2f}%" if hasattr(score, 'percentage') and getattr(score, 'percentage', None) else '',
                                '类型': getattr(score, 'type', '') if hasattr(score, 'type') else '',
                                '备注': getattr(score, 'comments', '') if hasattr(score, 'comments') else '',
                                '记录时间': getattr(score, 'recorded_at', '').strftime('%Y-%m-%d %H:%M:%S') if hasattr(score, 'recorded_at') and getattr(score, 'recorded_at') else ''
                            }
                            data.append(row_data)
                    except Exception as e:
                        print(f'[ExcelHandler] 转换第 {idx + 1} 条数据时出错: {e}')
                        import traceback
                        traceback.print_exc()
                        continue
                
                print(f'[ExcelHandler] 转换完成，共 {len(data)} 条数据')
                df = pd.DataFrame(data)
                print(f'[ExcelHandler] DataFrame 创建成功，行数: {len(df)}, 列数: {len(df.columns)}')
                if len(df) > 0:
                    print(f'[ExcelHandler] DataFrame 前几行:\n{df.head()}')
                else:
                    print('[ExcelHandler] 警告：DataFrame 为空！')
            
            # 创建Excel文件
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='成绩表')
                
                # 获取工作表并调整列宽
                worksheet = writer.sheets['成绩表']
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                        len(str(col))
                    ) + 2
                    col_letter = chr(65 + idx) if idx < 26 else chr(65 + idx // 26 - 1) + chr(65 + idx % 26)
                    worksheet.column_dimensions[col_letter].width = min(max_length, 50)
            
            output.seek(0)
            return output
            
        except Exception as e:
            # 如果出错，返回一个包含错误信息的空文件
            output = BytesIO()
            df = pd.DataFrame({'错误': [f'导出失败: {str(e)}']})
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)
            return output
    
    @staticmethod
    def generate_import_template():
        """
        生成成绩导入模板Excel文件
        
        Returns:
            BytesIO: Excel模板文件的字节流
        """
        try:
            # 创建模板数据（包含示例行和说明）
            template_data = {
                '学号': ['S000001', 'S000002', 'S000003', ''],
                '班级ID': [1, 1, 2, ''],
                '科目': ['数学', '数学', '英语', ''],
                '成绩': [85.5, 92.0, 78.5, ''],
                '总分': [100.0, 100.0, 100.0, ''],
                '考试ID': ['', '', '', '（可选）'],
                '备注': ['期中考试', '期中考试', '期中考试', '（可选）']
            }
            
            df = pd.DataFrame(template_data)
            
            # 创建Excel文件
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='成绩导入模板')
                
                # 获取工作表
                worksheet = writer.sheets['成绩导入模板']
                
                # 设置列宽
                column_widths = {
                    'A': 15,  # 学号
                    'B': 12,  # 班级ID
                    'C': 15,  # 科目
                    'D': 12,  # 成绩
                    'E': 12,  # 总分
                    'F': 12,  # 考试ID
                    'G': 20   # 备注
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 添加说明行（在标题行下方插入）
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    # 如果 openpyxl 不可用，跳过样式设置
                    Font = PatternFill = Alignment = None
                
                # 设置标题行样式
                if Font and PatternFill and Alignment:
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 在最后一行添加说明
                    last_row = len(df) + 1
                    worksheet.merge_cells(f'A{last_row + 1}:G{last_row + 1}')
                    note_cell = worksheet[f'A{last_row + 1}']
                    note_cell.value = '说明：1. 学号：支持学生ID或学号（system_account）；2. 班级ID：必须是您管理的班级ID；3. 科目、成绩、总分为必填项；4. 考试ID和备注为可选项；5. 请删除示例数据后填写实际数据'
                    note_cell.font = Font(size=9, color='666666', italic=True)
                    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    note_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # 设置数据行样式
                    for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # 如果样式不可用，至少添加说明文本
                    last_row = len(df) + 1
                    worksheet.merge_cells(f'A{last_row + 1}:G{last_row + 1}')
                    note_cell = worksheet[f'A{last_row + 1}']
                    note_cell.value = '说明：1. 学号：支持学生ID或学号（system_account）；2. 班级ID：必须是您管理的班级ID；3. 科目、成绩、总分为必填项；4. 考试ID和备注为可选项；5. 请删除示例数据后填写实际数据'
            
            output.seek(0)
            return output
            
        except Exception as e:
            # 如果出错，返回一个简单的模板
            output = BytesIO()
            df = pd.DataFrame({
                '学号': ['S000001'],
                '班级ID': [1],
                '科目': ['数学'],
                '成绩': [85.5],
                '总分': [100.0],
                '考试ID': [''],
                '备注': ['']
            })
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)
            return output
    
    @staticmethod
    def validate_excel_format(file):
        """
        验证Excel文件格式
        
        Args:
            file: Flask上传的文件对象
            
        Returns:
            tuple: (is_valid, error_message)
        """
        try:
            # 检查文件扩展名
            if not file.filename.endswith(('.xlsx', '.xls')):
                return False, '只支持Excel文件格式 (.xlsx, .xls)'
            
            # 尝试读取文件
            df = pd.read_excel(file)
            
            # 检查必需的列（支持中英文列名）
            column_mapping = {
                'student_id': ['student_id', '学号', '学生id', '学生ID', 'studentid', 'studentId'],
                'class_id': ['class_id', '班级id', '班级ID', '班级编号', 'classid', 'classId'],
                'subject': ['subject', '科目', '课程', '课程名称'],
                'score': ['score', '成绩', '得分', '分数'],
                'total_score': ['total_score', '总分', '满分', 'totalscore', 'totalScore']
            }
            
            normalized_columns = {}
            for standard_name, possible_names in column_mapping.items():
                for col in df.columns:
                    col_str = str(col).strip()
                    if col_str in possible_names or col_str.lower() == standard_name.lower():
                        normalized_columns[standard_name] = col
                        break
            
            required_columns = ['student_id', 'class_id', 'subject', 'score', 'total_score']
            missing_columns = [col for col in required_columns if col not in normalized_columns]
            
            if missing_columns:
                return False, f'缺少必需的列: {", ".join(missing_columns)}'
            
            return True, None
            
        except Exception as e:
            return False, f'文件格式错误: {str(e)}'
